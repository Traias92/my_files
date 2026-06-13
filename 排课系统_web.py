import streamlit as st
import random
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import pandas as pd
import os
import base64

# ------------------------ 默认配置 ------------------------
DEFAULT_SUBJECTS = [
    '语文', '数学', '英语', '道德与法治', '科学',
    '音乐', '美术', '体育与健康', '劳动', '综合实践',
    '校本课程', '地方课程', '信息科技'
]
DEFAULT_TEACHERS = [
    '殷金利', '武俊英', '刘梦', '卢新颖', '王玉红', '崔金金', '王然', '郭继茹',
    '杨文凤', '王引娣', '王秋月', '董红杰', '崔志敏', '高桂香', '毕春茹',
    '苑薇薇', '高喜超', '孙雅静', '李亚', '刘莹', '彭晓颖', '赵彦青', '连懿德'
]

# ------------------------ 核心函数 ------------------------
def build_teacher_tasks(arrangements):
    teacher_tasks = defaultdict(list)
    class_subject_hours = defaultdict(lambda: defaultdict(int))
    for item in arrangements:
        t = item['teacher']
        cid = item['class_id']
        sub = item['subject']
        h = item['hours']
        teacher_tasks[t].append((cid, sub, h))
        class_subject_hours[cid][sub] += h
    return dict(teacher_tasks), dict(class_subject_hours)

def generate_schedule(teacher_tasks, class_list, days, am, pm):
    total_periods = am + pm
    time_slots = [(d, p) for d in range(days) for p in range(total_periods)]
    main_subjects = {'语文', '数学', '英语'}

    tasks = []
    for teacher, arr in teacher_tasks.items():
        for cid, subj, hours in arr:
            for _ in range(hours):
                tasks.append((cid, subj, teacher))

    best_schedule = None
    best_unscheduled = None
    best_unscheduled_count = float('inf')

    for _ in range(20):
        class_occupied = {cid: set() for cid in class_list}
        teacher_occupied = {t: set() for t in teacher_tasks}
        schedule = {cid: {} for cid in class_list}
        unscheduled = []
        shuffled = tasks.copy()
        random.shuffle(shuffled)

        for cid, subj, teacher in shuffled:
            available = [slot for slot in time_slots
                         if slot not in class_occupied[cid]
                         and slot not in teacher_occupied[teacher]]
            if not available:
                unscheduled.append((cid, subj, teacher, 1))
                continue

            def priority(slot):
                day, period = slot
                score = 0
                if subj in main_subjects:
                    if period < 2 and period < am:
                        score += 10
                    elif period < am:
                        score += 5
                teacher_day = sum(1 for s in teacher_occupied[teacher] if s[0] == day)
                score -= teacher_day * 2
                if period >= 2:
                    prev1 = schedule[cid].get((day, period-1))
                    prev2 = schedule[cid].get((day, period-2))
                    if prev1 and prev1[0] == subj and prev2 and prev2[0] == subj:
                        score -= 3
                return score

            available.sort(key=priority, reverse=True)
            chosen = available[0]
            schedule[cid][chosen] = (subj, teacher)
            class_occupied[cid].add(chosen)
            teacher_occupied[teacher].add(chosen)

        if len(unscheduled) < best_unscheduled_count:
            best_unscheduled_count = len(unscheduled)
            best_schedule = schedule
            best_unscheduled = unscheduled
            if best_unscheduled_count == 0:
                break

    return best_schedule, best_unscheduled

def export_excel(schedule, class_list, days, am, pm):
    wb = Workbook()
    wb.remove(wb.active)
    total_periods = am + pm
    day_names = ["周一", "周二", "周三", "周四", "周五"][:days]
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold = Font(bold=True)
    sorted_cls = sorted(class_list, key=lambda x: (int(x.split('-')[0]), int(x.split('-')[1])))

    # 各班表
    for cid in sorted_cls:
        grade, cls = cid.split('-')
        ws = wb.create_sheet(f"{grade}年级{cls}班")
        ws.cell(1,1,"节次").font = bold; ws.cell(1,1).alignment=center; ws.cell(1,1).border=thin
        for i,dn in enumerate(day_names):
            c = ws.cell(1,i+2,dn); c.font=bold; c.alignment=center; c.border=thin
        for p in range(total_periods):
            row = p+2
            label = f"上午{p+1}" if p<am else f"下午{p-am+1}"
            c = ws.cell(row,1,label); c.alignment=center; c.border=thin; c.font=bold
            for d in range(days):
                entry = schedule.get(cid,{}).get((d,p))
                txt = f"{entry[0]}\n{entry[1]}" if entry else ""
                c = ws.cell(row,d+2,txt); c.alignment=center; c.border=thin
        ws.column_dimensions['A'].width = 8
        for i in range(2, days+2): ws.column_dimensions[get_column_letter(i)].width = 14

    # 总课表
    block = 1 + days
    for has_teacher, title in [(True,"全校总课表(含教师)"), (False,"全校总课表(仅科目)")]:
        ws = wb.create_sheet(title)
        for idx,cid in enumerate(sorted_cls):
            start = idx*block+1
            g,c = cid.split('-')
            ws.cell(1,start,f"{g}年级{c}班").font=bold; ws.cell(1,start).alignment=center; ws.cell(1,start).border=thin
            if block>1: ws.merge_cells(start_row=1,start_column=start,end_row=1,end_column=start+block-1)
        for idx,cid in enumerate(sorted_cls):
            start = idx*block+1
            ws.cell(2,start,"节次").font=bold; ws.cell(2,start).alignment=center; ws.cell(2,start).border=thin
            for d in range(days):
                c = ws.cell(2,start+1+d,day_names[d]); c.font=bold; c.alignment=center; c.border=thin
        for p in range(total_periods):
            row = p+3
            label = f"上午{p+1}" if p<am else f"下午{p-am+1}"
            for idx,cid in enumerate(sorted_cls):
                start = idx*block+1
                ws.cell(row,start,label).alignment=center; ws.cell(row,start).border=thin; ws.cell(row,start).font=bold
                for d in range(days):
                    entry = schedule.get(cid,{}).get((d,p))
                    txt = f"{entry[0]}\n{entry[1]}" if (entry and has_teacher) else (entry[0] if entry else "")
                    c = ws.cell(row,start+1+d,txt); c.alignment=center; c.border=thin
        for idx in range(len(sorted_cls)):
            start = idx*block+1
            ws.column_dimensions[get_column_letter(start)].width = 8
            for d in range(days): ws.column_dimensions[get_column_letter(start+1+d)].width = 14

    os.makedirs("课表输出", exist_ok=True)
    fname = f"课表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join("课表输出", fname)
    wb.save(path)
    return path

def get_download_link(path):
    with open(path, "rb") as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()
    return f'<a href="data:application/octet-stream;base64,{b64}" download="{os.path.basename(path)}">📥 下载课表文件</a>'

def parse_batch(text, class_list):
    new = []
    warns = []
    for line in text.strip().splitlines():
        if '：' not in line: continue
        teacher, _, rest = line.partition('：')
        teacher = teacher.strip()
        if not teacher: continue
        items = re.split(r'[，,;；]', rest)
        for item in items:
            item = item.strip()
            m = re.match(r'(\d+-\d+)', item)
            if not m: continue
            cid = m.group(1)
            if cid not in class_list:
                warns.append(f"班级{cid}不在列表中")
                continue
            rest_part = item[m.end():].strip()
            m2 = re.match(r'([^\d]+?)\s*[-]?\s*(\d+)\s*节?', rest_part)
            if not m2: continue
            sub = m2.group(1).strip()
            h = int(m2.group(2))
            if h<=0: continue
            new.append({'teacher':teacher, 'class_id':cid, 'subject':sub, 'hours':h})
    return new, warns

# ------------------------ 界面 ------------------------
st.set_page_config(page_title="排课系统", layout="wide")
st.title("📚 小学排课系统")

# 初始化状态
if 'arrangements' not in st.session_state:
    st.session_state.arrangements = []

# 学校设置
with st.expander("⚙️ 学校基本设置", expanded=True):
    col1, col2, col3 = st.columns(3)
    grade_count = col1.number_input("年级数", 1, 12, 6)
    class_counts_str = col2.text_input("每班数(逗号分隔)", "1,1,2,2,1,1")
    am = col1.number_input("上午节数", 1, 6, 3)
    pm = col2.number_input("下午节数", 1, 6, 3)
    days = col3.number_input("每周天数", 1, 7, 5)

    # 解析班级
    try:
        counts = [int(x.strip()) for x in class_counts_str.split(',')]
        if len(counts) != grade_count:
            st.error("班级数个数与年级数不匹配")
            class_list = []
        else:
            class_list = []
            for g in range(1, grade_count+1):
                for c in range(1, counts[g-1]+1):
                    class_list.append(f"{g}-{c}")
            st.success(f"班级列表：{', '.join(class_list)}")
    except:
        st.error("班级格式错误")
        class_list = []

st.divider()

# 添加方式选择
tab1, tab2 = st.tabs(["📝 逐条添加", "📋 批量导入"])

with tab1:
    with st.form("single_form", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns([3,2,3,2])
        teacher = c1.selectbox("教师", DEFAULT_TEACHERS, key="s_teacher")
        if class_list:
            cid = c2.selectbox("班级", class_list, key="s_class")
        else:
            cid = c2.text_input("班级", "1-1", key="s_class")
        subject = c3.selectbox("科目", DEFAULT_SUBJECTS, key="s_subject")
        hours = c4.number_input("节数", 1, 15, 1, key="s_hours")
        if st.form_submit_button("✅ 添加这一条"):
            if class_list and cid not in class_list:
                st.error("班级无效")
            else:
                st.session_state.arrangements.append({
                    'teacher': teacher,
                    'class_id': cid,
                    'subject': subject,
                    'hours': hours
                })
                st.success(f"已添加：{teacher} {cid} {subject} {hours}节")

with tab2:
    st.markdown("粘贴文本格式（教师：班级 科目-节数, ...），每行一位老师")
    batch_text = st.text_area("粘贴内容", height=150,
        placeholder="殷金利：1-1 语文-8, 1-1 校本课程-1\n杨文凤：1-1 数学-7, 1-1 劳动-1\n...")
    if st.button("📥 批量导入"):
        if not class_list:
            st.error("请先设置正确的班级列表")
        else:
            new_items, warns = parse_batch(batch_text, class_list)
            if not new_items:
                st.warning("未能解析到有效安排")
            else:
                st.session_state.arrangements.extend(new_items)
                st.success(f"成功导入 {len(new_items)} 条")
                if warns:
                    st.warning("\n".join(warns[:5]))

st.divider()

# 已添加清单
st.subheader("📌 已添加的课程安排")
if st.session_state.arrangements:
    df = pd.DataFrame(st.session_state.arrangements)
    st.dataframe(df, use_container_width=True)
    if st.button("🗑️ 清空全部安排"):
        st.session_state.arrangements = []
        st.experimental_rerun()
else:
    st.info("暂无安排")

st.divider()

# 生成课表
if st.button("🚀 生成课表 Excel", type="primary", use_container_width=True):
    if not class_list:
        st.error("班级列表无效")
    elif not st.session_state.arrangements:
        st.error("请先添加安排")
    else:
        with st.spinner("正在排课，请稍候..."):
            teacher_tasks, class_subjects = build_teacher_tasks(st.session_state.arrangements)
            # 检查超时
            total_slots = days * (am + pm)
            overload = []
            for cid in class_list:
                total_h = sum(class_subjects.get(cid, {}).values())
                if total_h > total_slots:
                    overload.append(f"{cid} 总课时{total_h} > 可用{total_slots}")
            if overload:
                st.warning("以下班级课时超标：" + "\n".join(overload))

            schedule, unscheduled = generate_schedule(teacher_tasks, class_list, days, am, pm)
            path = export_excel(schedule, class_list, days, am, pm)

            st.success("课表生成成功！")
            st.markdown(get_download_link(path), unsafe_allow_html=True)
            if unscheduled:
                st.warning(f"⚠️ 有 {len(unscheduled)} 节未排入（前10条）：")
                st.write("\n".join([f"{u[0]} {u[1]}({u[2]})" for u in unscheduled[:10]]))